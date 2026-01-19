"""
Excel to JSON Converter for OCMS Data Dictionary
Converts OCMS_Data_Dictionary.xlsx to JSON format matching docs/data-dictionary/*.json
"""

import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Any


def clean_value(value):
    """Clean cell values - convert None to empty string, strip whitespace"""
    if value is None:
        return ""
    return str(value).strip()


def to_boolean(value, field_name: str) -> bool:
    """Convert Y/N or Yes/No to boolean"""
    if not value:
        return False
    val = str(value).upper()
    if val in ["Y", "YES"]:
        return True
    if val in ["N", "NO"]:
        return False
    return False


def process_sheet(ws, sheet_name: str) -> Dict[str, Any]:
    """
    Process a worksheet and convert to structured JSON

    Returns structure matching docs/data-dictionary/*.json:
    {
        "table_name": {
            "description": "...",
            "columns": [
                {
                    "name": "...",
                    "type": "...",
                    "primary_key": true/false,
                    "nullable": true/false,
                    "default": null,
                    "description": "...",
                    "epic": "..."
                }
            ]
        }
    }
    """
    result = {}

    # Get header row
    headers = []
    for cell in ws[1]:
        headers.append(clean_value(cell.value))

    print(f"Processing sheet: {sheet_name}")

    current_table = None

    # Start from row 2 (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Extract key columns
        table_name = clean_value(row[0])
        table_desc = clean_value(row[1])
        primary_key = clean_value(row[2])
        column_name = clean_value(row[3])
        col_type = clean_value(row[4])
        default_value = clean_value(row[5])
        is_null = clean_value(row[6])
        description = clean_value(row[7])

        # Handle different column names for Epic (sheet dependent)
        epic = ""
        if len(row) > 8:
            epic = clean_value(row[8])

        # Skip empty rows
        if not table_name and not column_name:
            continue

        # Update current table if table_name exists
        if table_name:
            current_table = table_name
            if current_table not in result:
                result[current_table] = {
                    "description": table_desc,
                    "columns": []
                }

        # Add column if exists
        if column_name and current_table:
            # Convert primary_key to boolean
            is_pk = to_boolean(primary_key, "primary_key")

            # Convert is_null to boolean ("Yes" = nullable, "No" = not nullable)
            nullable = to_boolean(is_null, "nullable")

            # Handle default value
            default = None
            if default_value and default_value.lower() not in ["n/a", "null", ""]:
                default = default_value

            column_data = {
                "name": column_name,
                "type": col_type,
                "primary_key": is_pk,
                "nullable": nullable,
                "default": default,
                "description": description
            }

            # Only add epic if it exists
            if epic and epic.lower() not in ["n/a", "null", ""]:
                column_data["epic"] = epic

            result[current_table]["columns"].append(column_data)

        if row_idx % 100 == 0:
            print(f"  Processed {row_idx} rows...")

    return result


def convert_excel_to_json(excel_path: str, output_dir: str):
    """
    Convert Excel file to JSON format (one file per sheet)

    Args:
        excel_path: Path to input Excel file
        output_dir: Path to output directory
    """
    print(f"Loading Excel file: {excel_path}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Process each sheet and write to separate file
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = process_sheet(ws, sheet_name)

        # Determine output filename (lowercase)
        output_filename = f"{sheet_name.lower()}.json"
        output_file = output_path / output_filename

        # Write to JSON file
        print(f"Writing to: {output_file}")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(sheet_data, f, indent=2, ensure_ascii=False)

        num_tables = len(sheet_data)
        total_columns = sum(len(t["columns"]) for t in sheet_data.values())
        print(f"Completed: {num_tables} tables, {total_columns} columns")
        print()

    print("=" * 60)
    print("Conversion completed successfully!")
    print(f"Output directory: {output_dir}")
    print("=" * 60)


def main():
    """Main entry point"""
    # Paths
    base_dir = Path(r"C:\MGG\Tech doc")
    excel_file = base_dir / "Data dictionary" / "OCMS_Data_Dictionary.xlsx"
    output_dir = base_dir / "docs" / "data-dictionary"

    print("=" * 60)
    print("Excel to JSON Converter - OCMS Data Dictionary")
    print("=" * 60)
    print()

    # Check if input file exists
    if not excel_file.exists():
        print(f"Error: Excel file not found: {excel_file}")
        return

    # Convert
    convert_excel_to_json(str(excel_file), str(output_dir))


if __name__ == "__main__":
    main()
